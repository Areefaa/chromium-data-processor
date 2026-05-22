"""
chromium_to_excel.py
Menggabungkan 12 file CSV percobaan chromium ke dalam satu file .xlsx
yang terformat rapi dan siap diunduh.
"""

import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Konfigurasi ──────────────────────────────────────────────────────────────
CSV_FOLDER   = "data/raw"          # folder berisi 12 CSV
OUTPUT_FILE  = "output/chromium_data.xlsx"
SUMMARY_TITLE = "Percobaan Chromium – Data Lengkap"

# Warna tema
COLOR_HEADER_BG  = "1F4E79"   # biru tua
COLOR_HEADER_FG  = "FFFFFF"   # putih
COLOR_SUBHDR_BG  = "D6E4F0"   # biru muda
COLOR_ALT_ROW    = "EBF5FB"   # biru sangat muda (baris selang-seling)
COLOR_SUMMARY_BG = "FFF9C4"   # kuning muda (ringkasan)

# ── Helper styles ─────────────────────────────────────────────────────────────
def hdr_style(cell, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, size=11):
    cell.font      = Font(name="Arial", bold=True, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_table_style(ws, data_rows, header_row, n_cols):
    """Terapkan border, alternating row color, dan lebar kolom otomatis."""
    for row in ws.iter_rows(min_row=header_row, max_row=header_row + data_rows,
                             min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = thin_border()

    # Alternating row color
    for r_idx in range(header_row + 1, header_row + data_rows + 1):
        if (r_idx - header_row) % 2 == 0:
            for c_idx in range(1, n_cols + 1):
                ws.cell(r_idx, c_idx).fill = PatternFill("solid", start_color=COLOR_ALT_ROW)

    # Auto-width kolom
    for col in ws.iter_cols(min_row=header_row, max_row=header_row + data_rows,
                              min_col=1, max_col=n_cols):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

# ── 1. Baca semua CSV ─────────────────────────────────────────────────────────
def load_csvs(folder: str) -> dict[str, pd.DataFrame]:
    pattern = os.path.join(folder, "*.csv")
    files   = sorted(glob.glob(pattern))
    if not files:
        raise FileNotFoundError(f"Tidak ada file CSV di '{folder}'")
    dfs = {}
    for fp in files:
        name = os.path.splitext(os.path.basename(fp))[0]

        # Baca raw untuk deteksi separator dan baris header data
        with open(fp, encoding="utf-8", errors="replace") as f:
            lines = f.readlines()

        # Cari baris yang berisi "Wavelength" sebagai awal data sesungguhnya
        start_idx = 0
        meta_rows = []
        for i, line in enumerate(lines):
            if "Wavelength" in line or (";" in line and line.count(";") >= 1
                                        and i > 0 and "Wavelength" in lines[i]):
                start_idx = i
                break
            meta_rows.append(line.strip())

        # Gabungkan metadata jadi satu baris info (simpan sebagai atribut)
        meta_info = " | ".join([m for m in meta_rows if m])

        # Baca data mulai dari baris Wavelength, separator ";"
        import io
        data_text = "".join(lines[start_idx:])
        df = pd.read_csv(
            io.StringIO(data_text),
            sep=";",
            decimal=",",
            on_bad_lines="skip",
            engine="python"
        )

        # Simpan metadata sebagai kolom tambahan di baris pertama (opsional)
        df.attrs["meta"] = meta_info
        df.attrs["source_name"] = name

        dfs[name] = df
        print(f"✔ '{name}': {len(df)} baris, kolom: {list(df.columns)}")

    print(f"\n✔ Total {len(dfs)} file CSV dimuat.")
    return dfs

# ── 2. Tulis setiap CSV ke sheet tersendiri ───────────────────────────────────
def write_data_sheets(wb, dfs: dict[str, pd.DataFrame]):
    for sheet_name, df in dfs.items():
        # Potong nama sheet max 31 karakter (batas Excel)
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(title=safe_name)

        # Judul sheet
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(1, 1, value=f"Data: {sheet_name}")
        hdr_style(title_cell, size=12)
        ws.row_dimensions[1].height = 22

        # Header kolom
        for c_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(2, c_idx, value=col_name)
            hdr_style(cell, bg=COLOR_SUBHDR_BG, fg="000000")

        # Data
        for r_idx, row in enumerate(df.itertuples(index=False), 3):
            for c_idx, val in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value=val)

        apply_table_style(ws, len(df), header_row=2, n_cols=len(df.columns))

        # Freeze pane di bawah header
        ws.freeze_panes = "A3"

# ── 3. Sheet Ringkasan ────────────────────────────────────────────────────────
def write_summary_sheet(wb, dfs: dict[str, pd.DataFrame]):
    ws = wb.create_sheet(title="Ringkasan", index=0)
    ws.sheet_view.showGridLines = False

    # Judul utama
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = SUMMARY_TITLE
    t.font      = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    t.fill      = PatternFill("solid", start_color=COLOR_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-header tabel ringkasan
    headers = ["No", "Nama File / Eksperimen", "Jumlah Baris",
               "Jumlah Kolom", "Kolom-Kolom", "Keterangan"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(3, c_idx, value=h)
        hdr_style(cell, bg=COLOR_SUBHDR_BG, fg="000000")

    # Isi ringkasan
    for i, (name, df) in enumerate(dfs.items(), 1):
        row = 3 + i
        ws.cell(row, 1, value=i)
        ws.cell(row, 2, value=name)
        ws.cell(row, 3, value=f"=COUNTA('{name[:31]}'!A3:A10000)")  # formula dinamis
        ws.cell(row, 4, value=len(df.columns))
        ws.cell(row, 5, value=", ".join(df.columns.tolist()))
        ws.cell(row, 6, value="—")

        # Alternating color
        if i % 2 == 0:
            for c in range(1, 7):
                ws.cell(row, c).fill = PatternFill("solid", start_color=COLOR_ALT_ROW)

    # Apply border & lebar kolom
    total_rows = len(dfs)
    apply_table_style(ws, total_rows, header_row=3, n_cols=6)

    # Lebar kolom manual
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 20

    ws.freeze_panes = "A4"

# ── 4. Sheet Gabungan (All Data) ──────────────────────────────────────────────
def write_combined_sheet(wb, dfs: dict[str, pd.DataFrame]):
    """Gabungkan semua CSV ke satu sheet dengan kolom penanda 'Sumber'."""
    frames = []
    for name, df in dfs.items():
        temp = df.copy()
        temp.insert(0, "Sumber", name)
        frames.append(temp)

    combined = pd.concat(frames, ignore_index=True)

    ws = wb.create_sheet(title="Semua Data")

    # Judul
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(combined.columns))
    tc = ws.cell(1, 1, value="Gabungan Semua Data Percobaan Chromium")
    hdr_style(tc, size=12)
    ws.row_dimensions[1].height = 22

    # Header
    for c_idx, col in enumerate(combined.columns, 1):
        cell = ws.cell(2, c_idx, value=col)
        hdr_style(cell, bg=COLOR_SUBHDR_BG, fg="000000")

    # Data
    for r_idx, row in enumerate(combined.itertuples(index=False), 3):
        for c_idx, val in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value=val)

    apply_table_style(ws, len(combined), header_row=2, n_cols=len(combined.columns))
    ws.freeze_panes = "A3"

    print(f"✔ Sheet 'Semua Data' → {len(combined)} baris.")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs("output", exist_ok=True)

    dfs = load_csvs(CSV_FOLDER)

    # Buat workbook baru dengan sheet default lalu hapus
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)   # hapus sheet kosong default

    write_summary_sheet(wb, dfs)
    write_data_sheets(wb, dfs)
    write_combined_sheet(wb, dfs)

    wb.save(OUTPUT_FILE)
    print(f"✔ File tersimpan: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
