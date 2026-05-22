"""
smoothing.py
Savitzky-Golay smoothing untuk dataset percobaan chromium.
Window: 30 points | Polynomial Order: 2
"""

import os
import glob
import io
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Konfigurasi ───────────────────────────────────────────────────────────────
CSV_FOLDER      = "data/raw"
OUTPUT_FILE     = "output/chromium_smoothed.xlsx"
WINDOW_LENGTH   = 30     # Points of Window (harus genap → otomatis +1 jika perlu)
POLY_ORDER      = 2      # Polynomial Order

COLOR_HDR_BG    = "1A3A5C"
COLOR_HDR_FG    = "FFFFFF"
COLOR_RAW_BG    = "D6E4F0"
COLOR_SMOOTH_BG = "D5F5E3"
COLOR_ALT       = "F2F9FF"

# ── Helper ────────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=COLOR_HDR_BG, fg=COLOR_HDR_FG, bold=True, size=10):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def style_data(cell, alt=False):
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = thin_border()
    if alt:
        cell.fill = PatternFill("solid", start_color=COLOR_ALT)

# ── Baca CSV (format laboratorium: separator ";", desimal ",") ────────────────
def load_csvs(folder: str) -> dict:
    files = sorted(glob.glob(os.path.join(folder, "*.csv")))
    if not files:
        raise FileNotFoundError(f"Tidak ada CSV di '{folder}'")
    dfs = {}
    for fp in files:
        name = os.path.splitext(os.path.basename(fp))[0]
        with open(fp, encoding="utf-8", errors="replace") as f:
            lines = f.readlines()

        # Cari baris awal data (mengandung "Wavelength")
        start_idx = 0
        for i, line in enumerate(lines):
            if "Wavelength" in line:
                start_idx = i
                break

        data_text = "".join(lines[start_idx:])
        df = pd.read_csv(
            io.StringIO(data_text),
            sep=";",
            decimal=",",
            on_bad_lines="skip",
            engine="python"
        )
        # Bersihkan nama kolom
        df.columns = [c.strip() for c in df.columns]
        # Ambil hanya kolom numerik Wavelength & Absorbance
        df = df.dropna(how="all").reset_index(drop=True)
        dfs[name] = df
        print(f"✔ '{name}': {len(df)} baris")
    return dfs

# ── Savitzky-Golay Smoothing ──────────────────────────────────────────────────
def apply_savgol(series: pd.Series, window: int, poly: int) -> pd.Series:
    """Terapkan Savitzky-Golay. Window harus ganjil dan > poly."""
    w = window if window % 2 == 1 else window + 1   # pastikan ganjil
    w = max(w, poly + 2)                             # pastikan > poly
    if len(series) < w:
        print(f"  ⚠ Data terlalu pendek ({len(series)} baris) untuk window {w}, skip smoothing.")
        return series
    return pd.Series(savgol_filter(series.values, window_length=w, polyorder=poly),
                     index=series.index)

# ── Tulis sheet per-eksperimen ────────────────────────────────────────────────
def write_sheet(wb: Workbook, name: str, df: pd.DataFrame, window: int, poly: int):
    safe = name[:28]
    ws   = wb.create_sheet(title=safe)

    # Deteksi kolom wavelength dan absorbance
    cols = df.columns.tolist()
    wav_col = next((c for c in cols if "wave" in c.lower()), cols[0])
    abs_col = next((c for c in cols if "abs"  in c.lower()), cols[1] if len(cols) > 1 else cols[0])

    wavelength  = pd.to_numeric(df[wav_col], errors="coerce")
    absorbance  = pd.to_numeric(df[abs_col], errors="coerce")
    valid_mask  = wavelength.notna() & absorbance.notna()
    wavelength  = wavelength[valid_mask].reset_index(drop=True)
    absorbance  = absorbance[valid_mask].reset_index(drop=True)
    smoothed    = apply_savgol(absorbance, window, poly)

    # ── Header sheet ──
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = f"Savitzky-Golay Smoothing — {name}"
    t.font      = Font(name="Calibri", bold=True, size=12, color=COLOR_HDR_FG)
    t.fill      = PatternFill("solid", start_color=COLOR_HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Info parameter
    ws["A2"] = f"Window: {WINDOW_LENGTH} pts (aktif: {WINDOW_LENGTH if WINDOW_LENGTH%2==1 else WINDOW_LENGTH+1})  |  Poly Order: {poly}  |  Data: {len(wavelength)} titik"
    ws["A2"].font      = Font(name="Calibri", italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 16

    # ── Header tabel ──
    headers = [wav_col, f"{abs_col} (Raw)", f"{abs_col} (Smoothed)", "Δ Absorbance"]
    header_bg = [COLOR_HDR_BG, COLOR_RAW_BG, COLOR_SMOOTH_BG, "E8DAEF"]
    header_fg = [COLOR_HDR_FG, "000000", "000000", "000000"]
    for c, (h, bg, fg) in enumerate(zip(headers, header_bg, header_fg), 1):
        cell = ws.cell(3, c, value=h)
        style_header(cell, bg=bg, fg=fg)
    ws.row_dimensions[3].height = 18

    # ── Hitung nilai maksimum ──
    raw_max_idx = absorbance.idxmax()
    sm_max_idx  = smoothed.idxmax()

    raw_max_abs = absorbance[raw_max_idx]
    raw_max_wav = wavelength[raw_max_idx]
    sm_max_abs  = smoothed[sm_max_idx]
    sm_max_wav  = wavelength[sm_max_idx]

    # ── Kotak info maksimum (di atas tabel, baris 3–4 kolom E–F) ──
    COLOR_MAX_RAW = "1A5276"    # biru tua
    COLOR_MAX_SM  = "145A32"    # hijau tua

    labels = [
        ("Max Raw",      raw_max_wav, raw_max_abs, COLOR_MAX_RAW),
        ("Max Smoothed", sm_max_wav,  sm_max_abs,  COLOR_MAX_SM),
    ]
    for offset, (label, wav, abs_val, color) in enumerate(labels):
        r_box = 3 + offset
        # Label
        lc = ws.cell(r_box, 5, value=label)
        lc.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        lc.fill      = PatternFill("solid", start_color=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = thin_border()
        # Wavelength
        wc = ws.cell(r_box, 6, value=round(wav, 4))
        wc.font      = Font(name="Calibri", bold=True, size=10, color=color)
        wc.alignment = Alignment(horizontal="right", vertical="center")
        wc.border    = thin_border()
        wc.number_format = "0.0000"
        # Absorbance
        ac = ws.cell(r_box, 7, value=round(abs_val, 9))
        ac.font      = Font(name="Calibri", bold=True, size=10, color=color)
        ac.alignment = Alignment(horizontal="right", vertical="center")
        ac.border    = thin_border()
        ac.number_format = "0.000000000"

    # Header kotak info
    for c_idx, hdr in enumerate(["Keterangan", "λ Maks (nm)", "Absorbansi Maks"], 5):
        hc = ws.cell(2, c_idx, value=hdr)
        style_header(hc, bg="2C3E50", fg="FFFFFF", size=9)
    ws.merge_cells("A2:D2")   # tetap merge A2:D2 untuk info parameter

    # ── Data ──
    for r, (wv, raw, sm) in enumerate(zip(wavelength, absorbance, smoothed), 4):
        alt = (r % 2 == 0)

        # Sorot baris nilai maksimum
        is_raw_max = (r - 4) == raw_max_idx
        is_sm_max  = (r - 4) == sm_max_idx

        cells = [
            ws.cell(r, 1, value=round(wv, 6)),
            ws.cell(r, 2, value=round(raw, 9)),
            ws.cell(r, 3, value=round(sm, 9)),
            ws.cell(r, 4, value=round(sm - raw, 9)),
        ]
        for cell in cells:
            style_data(cell, alt=alt)

        # Override warna baris max raw → biru muda terang
        if is_raw_max:
            for cell in cells[:2]:
                cell.fill = PatternFill("solid", start_color="AED6F1")
                cell.font = Font(name="Calibri", bold=True, size=10, color="1A5276")
        # Override warna baris max smoothed → hijau muda terang
        if is_sm_max:
            for cell in [cells[0], cells[2]]:
                cell.fill = PatternFill("solid", start_color="A9DFBF")
                cell.font = Font(name="Calibri", bold=True, size=10, color="145A32")

    # Format angka
    for r in range(4, 4 + len(wavelength)):
        ws.cell(r, 1).number_format = "0.0000"
        ws.cell(r, 2).number_format = "0.000000000"
        ws.cell(r, 3).number_format = "0.000000000"
        ws.cell(r, 4).number_format = "0.000000000"

    # Lebar kolom
    widths = [16, 20, 20, 18, 16, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    # ── Grafik perbandingan Raw vs Smoothed ──
    n_rows = len(wavelength)
    chart  = LineChart()
    chart.title  = f"Raw vs Smoothed — {name[:25]}"
    chart.style  = 10
    chart.y_axis.title = "Absorbance"
    chart.x_axis.title = "Wavelength (nm)"
    chart.height = 12
    chart.width  = 22

    # Raw
    raw_ref  = Reference(ws, min_col=2, min_row=3, max_row=3 + n_rows)
    chart.add_data(raw_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill  = "5B9BD5"
    chart.series[0].graphicalProperties.line.width      = 10000   # 0.75pt
    chart.series[0].smooth = False

    # Smoothed
    sm_ref = Reference(ws, min_col=3, min_row=3, max_row=3 + n_rows)
    chart.add_data(sm_ref, titles_from_data=True)
    chart.series[1].graphicalProperties.line.solidFill  = "E74C3C"
    chart.series[1].graphicalProperties.line.width      = 20000   # 1.5pt
    chart.series[1].smooth = True

    # X axis = wavelength
    x_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + n_rows)
    chart.set_categories(x_ref)

    ws.add_chart(chart, "F3")

    return {
        "name": name, "n_points": n_rows, "wav_col": wav_col, "abs_col": abs_col,
        "raw_max_wav": round(raw_max_wav, 4), "raw_max_abs": round(raw_max_abs, 9),
        "sm_max_wav":  round(sm_max_wav,  4), "sm_max_abs":  round(sm_max_abs,  9),
    }

# ── Sheet Ringkasan ───────────────────────────────────────────────────────────
def write_summary(wb: Workbook, results: list):
    ws = wb.create_sheet(title="Ringkasan", index=0)

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "Ringkasan Savitzky-Golay Smoothing — Percobaan Chromium"
    t.font      = Font(name="Calibri", bold=True, size=13, color=COLOR_HDR_FG)
    t.fill      = PatternFill("solid", start_color=COLOR_HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Parameter box
    params = [
        ("Metode", "Savitzky-Golay"),
        ("Points of Window", f"{WINDOW_LENGTH} (aktif: {WINDOW_LENGTH if WINDOW_LENGTH%2==1 else WINDOW_LENGTH+1})"),
        ("Polynomial Order", str(POLY_ORDER)),
        ("Jumlah File", str(len(results))),
    ]
    for i, (k, v) in enumerate(params, 2):
        ws.cell(i, 1, value=k).font  = Font(name="Calibri", bold=True, size=10)
        ws.cell(i, 2, value=v).font  = Font(name="Calibri", size=10)
        ws.row_dimensions[i].height  = 16

    # Tabel ringkasan
    start_row = len(params) + 3
    hdrs = [
        "No", "Nama Eksperimen", "Jumlah Titik",
        "λ Maks Raw (nm)", "Abs Maks Raw",
        "λ Maks Smoothed (nm)", "Abs Maks Smoothed",
    ]
    hdr_colors = [
        COLOR_HDR_BG, COLOR_HDR_BG, COLOR_HDR_BG,
        "1A5276", "1A5276",
        "145A32", "145A32",
    ]
    for c, (h, bg) in enumerate(zip(hdrs, hdr_colors), 1):
        style_header(ws.cell(start_row, c, value=h), bg=bg)

    for i, r in enumerate(results, 1):
        row = start_row + i
        alt = i % 2 == 0
        data = [
            i, r["name"], r["n_points"],
            r["raw_max_wav"], r["raw_max_abs"],
            r["sm_max_wav"],  r["sm_max_abs"],
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row, c, value=val)
            style_data(cell, alt=alt)
            if c in (4, 5):
                cell.number_format = "0.0000" if c == 4 else "0.000000000"
                cell.font = Font(name="Calibri", bold=True, color="1A5276", size=10)
            if c in (6, 7):
                cell.number_format = "0.0000" if c == 6 else "0.000000000"
                cell.font = Font(name="Calibri", bold=True, color="145A32", size=10)

    widths = [5, 32, 14, 20, 22, 22, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs("output", exist_ok=True)

    dfs     = load_csvs(CSV_FOLDER)
    wb      = Workbook()
    wb.remove(wb.active)

    results = []
    for name, df in dfs.items():
        print(f"  → Smoothing '{name}' ...")
        info = write_sheet(wb, name, df, WINDOW_LENGTH, POLY_ORDER)
        results.append(info)

    write_summary(wb, results)
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Selesai! File tersimpan: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
