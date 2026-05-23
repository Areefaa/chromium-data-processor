"""
analisis_penurunan.py
Membaca data spektra dari file Excel yang sudah ada, lalu menghasilkan
file baru "Data_Penelitian_Beserta_Spektra.xlsx" dengan sheet analisis
penurunan absorbansi lengkap.

Rumus:
    Penurunan absorbansi (%) = (A0 - At) / A0 × 100

A0 = absorbansi maksimum smoothed FP_5000X (baseline, tetap)
At = absorbansi maksimum smoothed tiap variasi
"""

import os
import io
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ── Konfigurasi ───────────────────────────────────────────────────────────────
INPUT_FILE    = "output/Data_Penelitian_Spektra.xlsx"  # ← sumber data
OUTPUT_FILE   = "output/Data_Penelitian_Beserta_Spektra.xlsx"  # ← hasil (timpa)
WINDOW_LENGTH = 30
POLY_ORDER    = 2
BASELINE_KEY  = "FP_5000X"   # ← nama sheet baseline (tanpa .csv)

# Sheet yang dilewati saat membaca (bukan sheet spektra)
SKIP_SHEETS   = {"Analisis Penurunan Absorbansi", "Ringkasan"}

# Warna
C_DARK_BLUE = "1A3A5C"
C_WHITE     = "FFFFFF"
C_ALT       = "F4F9FF"
C_A0        = "154360"
C_PCT_POS   = "1E8449"
C_PCT_NEG   = "C0392B"
C_GOLD      = "B7950B"

# ── Helpers ───────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=C_DARK_BLUE, fg=C_WHITE, size=10, bold=True):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def style_data(cell, alt=False, align="right"):
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = thin_border()
    if alt:
        cell.fill = PatternFill("solid", start_color=C_ALT)

# ── Baca dari Excel ───────────────────────────────────────────────────────────
def load_from_excel(filepath: str) -> dict:
    """
    Baca tiap sheet spektra dari file Excel yang sudah ada.
    Skip sheet ringkasan/analisis.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(
            f"\n❌ File '{filepath}' tidak ditemukan!\n"
            f"   Pastikan file sudah ada di folder output/"
        )

    wb   = load_workbook(filepath, data_only=True)
    dfs  = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws   = wb[sheet_name]
        rows = list(ws.values)

        # Cari baris header (mengandung kata "Wavelength")
        header_idx = None
        for i, row in enumerate(rows):
            if row and any("wavelength" in str(v or "").lower() for v in row):
                header_idx = i
                break

        if header_idx is None:
            print(f"  ⚠ Sheet '{sheet_name}': header Wavelength tidak ditemukan, dilewati.")
            continue

        headers = [str(v or "").strip() for v in rows[header_idx]]
        data    = [r for r in rows[header_idx + 1:] if any(v is not None for v in r)]
        df      = pd.DataFrame(data, columns=headers)
        df      = df.dropna(how="all").reset_index(drop=True)

        # Bersihkan kolom — ambil hanya Wavelength & Absorbance (2 kolom pertama numerik)
        numeric_cols = []
        for col in df.columns:
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().sum() > 10:   # minimal 10 baris valid
                numeric_cols.append(col)
            if len(numeric_cols) == 2:
                break

        if len(numeric_cols) < 2:
            print(f"  ⚠ Sheet '{sheet_name}': kolom numerik tidak cukup, dilewati.")
            continue

        df = df[numeric_cols].copy()
        dfs[sheet_name] = df
        print(f"  ✔ '{sheet_name}': {len(df)} baris  |  kolom: {numeric_cols}")

    wb.close()
    return dfs

# ── Savitzky-Golay ────────────────────────────────────────────────────────────
def apply_savgol(series: pd.Series, window: int, poly: int) -> pd.Series:
    w = window if window % 2 == 1 else window + 1
    w = max(w, poly + 2)
    if len(series) < w:
        return series
    return pd.Series(
        savgol_filter(series.values, window_length=w, polyorder=poly),
        index=series.index
    )

# ── Ekstrak spektra + hitung nilai maks ──────────────────────────────────────
def extract_spectrum(df: pd.DataFrame) -> dict:
    cols    = df.columns.tolist()
    wav_col = next((c for c in cols if "wave" in c.lower()), cols[0])
    abs_col = next((c for c in cols if "abs"  in c.lower()),
                   cols[1] if len(cols) > 1 else cols[0])

    wavelength = pd.to_numeric(df[wav_col], errors="coerce")
    absorbance = pd.to_numeric(df[abs_col], errors="coerce")
    mask       = wavelength.notna() & absorbance.notna()
    wavelength = wavelength[mask].reset_index(drop=True)
    absorbance = absorbance[mask].reset_index(drop=True)
    smoothed   = apply_savgol(absorbance, WINDOW_LENGTH, POLY_ORDER)

    sm_max_idx = smoothed.idxmax()
    return {
        "wav_col":    wav_col,
        "abs_col":    abs_col,
        "wavelength": wavelength,
        "absorbance": absorbance,
        "smoothed":   smoothed,
        "sm_max_abs": float(smoothed[sm_max_idx]),
        "sm_max_wav": float(wavelength[sm_max_idx]),
    }

# ── Sheet Analisis Penurunan Absorbansi ───────────────────────────────────────
def write_analisis_sheet(wb: Workbook, spectra: dict, A0: float, baseline_key: str):
    ws = wb.create_sheet(title="Analisis Penurunan Absorbansi", index=0)
    ws.sheet_view.showGridLines = False

    # Judul utama
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "Analisis Penurunan Absorbansi — Percobaan Chromium"
    t.font      = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    t.fill      = PatternFill("solid", start_color=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Kotak rumus
    ws.merge_cells("A2:F2")
    r2 = ws["A2"]
    r2.value     = "Rumus:  Penurunan absorbansi (%) = (A₀ − Aₜ) / A₀ × 100"
    r2.font      = Font(name="Calibri", italic=True, size=11, color="2C3E50")
    r2.alignment = Alignment(horizontal="center", vertical="center")
    r2.fill      = PatternFill("solid", start_color="EBF5FB")
    ws.row_dimensions[2].height = 20

    # Keterangan variabel
    ws.merge_cells("A3:F3")
    r3 = ws["A3"]
    r3.value = (f"A₀ = Abs maks smoothed {baseline_key} (baseline, tetap)"
                f"    |    Aₜ = Abs maks smoothed tiap variasi (setelah perlakuan)"
                f"    |    A₀ = {round(A0, 9)}")
    r3.font      = Font(name="Calibri", italic=True, size=9, color="566573")
    r3.alignment = Alignment(horizontal="center", vertical="center")
    r3.fill      = PatternFill("solid", start_color="F2F3F4")
    ws.row_dimensions[3].height = 15

    ws.row_dimensions[4].height = 6   # pemisah

    # Header tabel
    hdrs   = ["No", "Nama Eksperimen", "λ Maks Smoothed (nm)",
              "Aₜ — Absorbansi Maks Smoothed", "Keterangan", "Penurunan Absorbansi (%)"]
    hdr_bg = [C_DARK_BLUE, C_DARK_BLUE, C_DARK_BLUE, C_DARK_BLUE, C_DARK_BLUE, C_GOLD]
    for c, (h, bg) in enumerate(zip(hdrs, hdr_bg), 1):
        cell = ws.cell(5, c, value=h)
        style_header(cell, bg=bg)
    ws.row_dimensions[5].height = 22

    names = list(spectra.keys())

    for i, name in enumerate(names, 1):
        sp          = spectra[name]
        At          = sp["sm_max_abs"]
        row         = 5 + i
        alt         = i % 2 == 0
        is_baseline = (name == baseline_key)

        if is_baseline:
            pct_val   = None
            pct_disp  = "—  (A₀ baseline)"
            ket_str   = f"A₀  ← {baseline_key}"
            row_color = "D6EAF8"
        else:
            pct_val   = (A0 - At) / A0 * 100 if A0 != 0 else 0
            pct_disp  = pct_val
            ket_str   = "Aₜ  (setelah perlakuan)"
            row_color = None

        cells_data = [
            (1, i),
            (2, name),
            (3, round(sp["sm_max_wav"], 4)),
            (4, round(At, 9)),
            (5, ket_str),
            (6, pct_disp),
        ]
        for c_idx, val in cells_data:
            cell = ws.cell(row, c_idx, value=val)
            style_data(cell, alt=alt,
                       align="center" if c_idx in (1, 5) else "right")
            if row_color:
                cell.fill = PatternFill("solid", start_color=row_color)
                cell.font = Font(name="Calibri", bold=True, size=10, color=C_A0)

        # Format angka
        ws.cell(row, 3).number_format = "0.0000"
        ws.cell(row, 4).number_format = "0.000000000"

        # Format & warna % penurunan
        pct_cell = ws.cell(row, 6)
        if pct_val is not None:
            pct_cell.number_format = "0.00"
            is_pos = pct_val >= 0
            pct_cell.font = Font(name="Calibri", bold=True, size=10,
                                 color=C_PCT_POS if is_pos else C_PCT_NEG)
            pct_cell.fill = PatternFill("solid",
                                        start_color="EAFAF1" if is_pos else "FDEDEC")

    # Lebar kolom & freeze
    ws.freeze_panes = "A6"
    for i, w in enumerate([5, 30, 22, 30, 26, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Grafik tren % penurunan
    n     = len(names)
    chart = LineChart()
    chart.title        = "Tren Penurunan Absorbansi (%)"
    chart.style        = 10
    chart.y_axis.title = "Penurunan (%)"
    chart.x_axis.title = "Eksperimen"
    chart.height       = 12
    chart.width        = 24

    data_ref = Reference(ws, min_col=6, min_row=5, max_row=5 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "1E8449"
    chart.series[0].graphicalProperties.line.width     = 20000
    chart.series[0].smooth = True

    cat_ref = Reference(ws, min_col=2, min_row=6, max_row=5 + n)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, "A" + str(8 + n))

# ── Sheet spektra per-eksperimen ──────────────────────────────────────────────
def write_spektra_sheet(wb: Workbook, name: str, sp: dict, A0: float, baseline_key: str):
    safe = name[:28]
    ws   = wb.create_sheet(title=safe)

    wav_col    = sp["wav_col"]
    abs_col    = sp["abs_col"]
    wavelength = sp["wavelength"]
    absorbance = sp["absorbance"]
    smoothed   = sp["smoothed"]
    sm_max_abs = sp["sm_max_abs"]
    sm_max_wav = sp["sm_max_wav"]

    is_baseline = (name == baseline_key)
    At  = sm_max_abs
    pct = 0.0 if is_baseline else ((A0 - At) / A0 * 100 if A0 != 0 else 0)

    # Header sheet
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = f"Spektra — {name}" + (" [A₀ BASELINE]" if is_baseline else "")
    t.font      = Font(name="Calibri", bold=True, size=12, color=C_WHITE)
    t.fill      = PatternFill("solid", start_color="154360" if is_baseline else C_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Info baris
    pct_info = "A₀ (Baseline — tidak dihitung penurunan)" if is_baseline \
               else f"Penurunan: {round(pct, 2)}%"
    info_str = (f"λ Maks Smoothed: {round(sm_max_wav, 4)} nm  |  "
                f"Abs Maks Smoothed: {round(sm_max_abs, 9)}  |  "
                f"{pct_info}  |  "
                f"S-G Window: {WINDOW_LENGTH} "
                f"(aktif: {WINDOW_LENGTH if WINDOW_LENGTH % 2 == 1 else WINDOW_LENGTH + 1})"
                f"  Poly: {POLY_ORDER}")
    ws.merge_cells("A2:E2")
    info = ws["A2"]
    info.value     = info_str
    info.font      = Font(name="Calibri", italic=True, size=9, color="555555")
    info.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # Header tabel
    headers = [wav_col, f"{abs_col} (Raw)", f"{abs_col} (Smoothed)", "Δ Abs",
               "% Penurunan*" if not is_baseline else "Keterangan"]
    hdr_bgs = [C_DARK_BLUE, "1A5276", "145A32", "6C3483",
               "7D6608" if not is_baseline else "154360"]
    for c, (h, bg) in enumerate(zip(headers, hdr_bgs), 1):
        style_header(ws.cell(3, c, value=h), bg=bg)
    ws.row_dimensions[3].height = 18

    sm_max_idx  = smoothed.idxmax()
    raw_max_idx = absorbance.idxmax()

    for r, (wv, raw, sm) in enumerate(zip(wavelength, absorbance, smoothed), 4):
        alt        = r % 2 == 0
        is_sm_max  = (r - 4) == sm_max_idx
        is_raw_max = (r - 4) == raw_max_idx
        row_pct    = 0.0 if is_baseline else (
                     (A0 - float(sm)) / A0 * 100 if A0 != 0 else 0)
        col5_val   = "A₀ Baseline" if is_baseline else round(row_pct, 4)

        vals = [round(wv, 6), round(raw, 9), round(sm, 9), round(sm - raw, 9), col5_val]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(r, c_idx, value=val)
            style_data(cell, alt=alt)

        if is_sm_max:
            for c_idx in range(1, 6):
                c = ws.cell(r, c_idx)
                c.fill = PatternFill("solid", start_color="A9DFBF")
                c.font = Font(name="Calibri", bold=True, size=10, color="145A32")
        elif is_raw_max:
            for c_idx in (1, 2):
                c = ws.cell(r, c_idx)
                c.fill = PatternFill("solid", start_color="AED6F1")
                c.font = Font(name="Calibri", bold=True, size=10, color="1A5276")

    # Format angka
    for r in range(4, 4 + len(wavelength)):
        ws.cell(r, 1).number_format = "0.0000"
        ws.cell(r, 2).number_format = "0.000000000"
        ws.cell(r, 3).number_format = "0.000000000"
        ws.cell(r, 4).number_format = "0.000000000"
        if not is_baseline:
            ws.cell(r, 5).number_format = "0.0000"

    # Catatan kaki
    note_row = 4 + len(wavelength) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    note_txt = (f"A₀ (baseline {baseline_key}) = {round(A0, 9)}"
                if not is_baseline else
                f"File ini adalah baseline (A₀). Nilai A₀ = {round(A0, 9)}")
    note = ws.cell(note_row, 1,
                   value=f"*) % Penurunan per-titik = (A₀ − A_smoothed) / A₀ × 100"
                         f"  |  {note_txt}")
    note.font      = Font(name="Calibri", italic=True, size=8, color="777777")
    note.alignment = Alignment(horizontal="left")

    for i, w in enumerate([16, 20, 20, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # Grafik Raw vs Smoothed
    n_rows = len(wavelength)
    chart  = LineChart()
    chart.title         = f"{'[A₀] ' if is_baseline else ''}Spektra — {name[:20]}"
    chart.style         = 10
    chart.y_axis.title  = "Absorbance"
    chart.x_axis.title  = "Wavelength (nm)"
    chart.height        = 11
    chart.width         = 20

    raw_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + n_rows)
    chart.add_data(raw_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "5B9BD5"
    chart.series[0].graphicalProperties.line.width     = 8000
    chart.series[0].smooth = False

    sm_ref = Reference(ws, min_col=3, min_row=3, max_row=3 + n_rows)
    chart.add_data(sm_ref, titles_from_data=True)
    chart.series[1].graphicalProperties.line.solidFill = "E74C3C"
    chart.series[1].graphicalProperties.line.width     = 18000
    chart.series[1].smooth = True

    x_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + n_rows)
    chart.set_categories(x_ref)
    ws.add_chart(chart, "G3")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs("output", exist_ok=True)

    print("── Membaca data dari Excel ──")
    print(f"   Sumber: {INPUT_FILE}\n")
    dfs = load_from_excel(INPUT_FILE)

    if not dfs:
        raise ValueError("❌ Tidak ada sheet spektra yang berhasil dibaca!")

    # Validasi baseline
    if BASELINE_KEY not in dfs:
        available = list(dfs.keys())
        raise ValueError(
            f"\n❌ Sheet baseline '{BASELINE_KEY}' tidak ditemukan!\n"
            f"   Sheet yang tersedia: {available}\n"
            f"   → Ubah nilai BASELINE_KEY di bagian konfigurasi atas."
        )

    print("\n── Ekstrak spektra & smoothing ──")
    spectra = {name: extract_spectrum(df) for name, df in dfs.items()}

    A0 = spectra[BASELINE_KEY]["sm_max_abs"]
    print(f"\n  ✔ A₀ (baseline '{BASELINE_KEY}') = {round(A0, 9)}")
    print(f"  ✔ Jumlah variasi (Aₜ)            : {len(spectra) - 1} eksperimen\n")

    wb = Workbook()
    wb.remove(wb.active)

    print("── Menulis sheet Analisis Penurunan ──")
    write_analisis_sheet(wb, spectra, A0, BASELINE_KEY)

    print("── Menulis sheet Spektra ──")
    for name, sp in spectra.items():
        print(f"  → '{name}' ...")
        write_spektra_sheet(wb, name, sp, A0, BASELINE_KEY)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Selesai! File tersimpan: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
